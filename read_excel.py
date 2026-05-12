import openpyxl
import json

src = r'D:\work\龙虾可操作需求临时文件夹\企业采购监管数据质检模型.xlsx'

wb = openpyxl.load_workbook(src)
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    for r in range(1, min(8, ws.max_row + 1)):
        row_data = {}
        for c in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            val = ws.cell(row=r, column=c).value
            row_data[col_letter] = str(val) if val is not None else ''
        print(f"Row {r}: {json.dumps(row_data, ensure_ascii=False)}")
