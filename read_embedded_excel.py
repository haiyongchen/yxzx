# -*- coding: utf-8 -*-
import openpyxl

excel_path = 'D:\\openclaw-workspace\\extracted_files\\word\\embeddings\\Workbook1.xlsx'
output_path = 'D:\\openclaw-workspace\\embedded_excel_content.txt'

# 读取Excel
wb = openpyxl.load_workbook(excel_path)

with open(output_path, 'w', encoding='utf-8') as f:
    f.write("=" * 80 + "\n")
    f.write("嵌入的Excel数据内容\n")
    f.write("=" * 80 + "\n")
    
    for sheet_name in wb.sheetnames:
        f.write(f"\nSheet: {sheet_name}\n")
        sheet = wb[sheet_name]
        
        # 获取数据
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            data.append(row)
        
        # 打印数据
        for i, row in enumerate(data):
            f.write(f"  行{i+1}: {row}\n")
    
    f.write("\n" + "=" * 80 + "\n")

print(f"Excel内容已保存到: {output_path}")
