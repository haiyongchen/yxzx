from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 读取现有文件
wb = load_workbook('技术标0分问题深度分析及AI评标规则库.xlsx')

# 创建新的Sheet：按评分点问题规律总结
ws_new = wb.create_sheet('按评分点规律总结', 1)

headers = ['评分点', '该评分点0分问题数', '主要问题类型', '高频问题描述', '问题占比', '核心规律', 'AI识别重点', '改进方向']
ws_new.append(headers)

# 按评分点整理的数据
score_point_analysis = [
    ['综合阐述', '8', 
     '针对性不足、内容偏离、信息错误', 
     '无针对本小区的具体内容(4次)、目标不清晰、范围错误、内容混乱', 
     '针对性问题占50%', 
     '缺乏对本项目的具体描述是最大问题，其次是内容偏离和基本信息错误', 
     '检测是否包含"本标段/本项目/本小区"、检测是否包含无关内容(融资/股权/危险化学品等)、检测目标是否明确', 
     '必须针对本项目展开，避免通用模板，确保基本信息准确'],
    
    ['管理架构', '5', 
     '针对性不足、描述不具体', 
     '针对本标段的安排欠缺/不详细(4次)、职责描述不具体', 
     '针对性问题占80%', 
     '组织架构的针对性安排是核心问题，其次是职责描述空泛', 
     '检测组织架构图正确性、检测是否针对本标段、检测职责描述具体程度', 
     '组织架构要正确，职责描述要具体，必须针对本标段'],
    
    ['施工进度计划', '7', 
     '图表缺失、描述不详细、阶段划分不清', 
     '无横道图/网络图(3次)、描述不详细(2次)、阶段划分不清(2次)', 
     '图表问题占43%', 
     '图表缺失是最严重问题，其次是图表说明不充分和时间节点不明确', 
     '检测横道图和网络图是否存在、检测图表是否有详细说明、检测是否标注日历天数', 
     '必须包含横道图和网络图，图表要有详细说明，时间节点要明确'],
    
    ['施工技术方案', '4', 
     '针对性不足、内容不全面、内容偏离', 
     '针对本标段的方案论述不全面(2次)、安全进度内容不全面、内容偏离', 
     '针对性问题占50%', 
     '方案论述不全面是主要问题，其次是内容偏离（如危险化学品）', 
     '检测是否针对本标段、检测安全/质量/进度三方面是否覆盖、检测是否有无关内容', 
     '方案必须针对本标段，覆盖安全/质量/进度三方面，避免无关内容'],
    
    ['风险源管控方案', '3', 
     '针对性差、风险识别不足', 
     '针对性差(2次)、针对本项目的风险欠缺', 
     '针对性问题占100%', 
     '所有0分原因都是针对性问题，风险识别不全面是核心', 
     '检测是否针对本项目识别风险、检测管控措施针对性', 
     '必须针对本项目识别具体风险，管控措施要有针对性'],
    
    ['关键部位质量管控措施', '3', 
     '针对性不足、措施质量差', 
     '针对项目的管控欠缺(2次)、措施质量差', 
     '针对性问题占67%', 
     '针对本项目的管控欠缺是主要问题，其次是措施本身质量不高', 
     '检测是否针对本项目、检测措施完善程度', 
     '必须针对本项目，措施要完善有效'],
    
    ['施工进度保证措施', '1', 
     '内容不详细', 
     '内容不够详细', 
     '内容问题占100%', 
     '内容简略是唯问题，但该评分点0分案例较少', 
     '检测内容详细程度', 
     '内容要详细具体'],
    
    ['质量安全文明施工保证措施', '1', 
     '针对性不足', 
     '针对本项目的措施欠缺', 
     '针对性问题占100%', 
     '针对性不足是唯一问题，该评分点0分案例较少', 
     '检测是否针对本项目', 
     '必须针对本项目编制措施'],
    
    ['扬尘污染治理方案', '1', 
     '措施不完善', 
     '降噪、残土排运措施差', 
     '措施问题占100%', 
     '具体措施不到位是唯一问题，该评分点0分案例较少', 
     '检测降噪/残土排运等措施完善程度', 
     '降噪、残土排运等措施要完善'],
    
    ['突发事件应急预案', '0', 
     '（无0分案例）', 
     '（无）', 
     '（无）', 
     '该评分点在样本数据中无0分案例，可能说明投标人普遍重视或标准较宽松', 
     '检测预案完整性、组织措施、抢救措施', 
     '确保预案完整，包含组织措施和抢救措施'],
]

for row_data in score_point_analysis:
    ws_new.append(row_data)

# 样式设置
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col in range(1, 9):
    cell = ws_new.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

for row in range(2, len(score_point_analysis) + 2):
    for col in range(1, 9):
        cell = ws_new.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = data_align

# 设置列宽
ws_new.column_dimensions['A'].width = 22
ws_new.column_dimensions['B'].width = 15
ws_new.column_dimensions['C'].width = 28
ws_new.column_dimensions['D'].width = 40
ws_new.column_dimensions['E'].width = 15
ws_new.column_dimensions['F'].width = 45
ws_new.column_dimensions['G'].width = 45
ws_new.column_dimensions['H'].width = 40

ws_new.row_dimensions[1].height = 40
for row in range(2, len(score_point_analysis) + 2):
    ws_new.row_dimensions[row].height = 60

# 保存
wb.save('D:/openclaw-workspace/技术标0分问题深度分析及AI评标规则库_v2.xlsx')
print('已添加按评分点规律总结Sheet')
