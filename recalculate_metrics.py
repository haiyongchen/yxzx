#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按照新指标定义重新统计
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 读取原始数据
file_path = r'D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\temp\专区信息汇总表_中原华北_合并.xlsx'
df = pd.read_excel(file_path)

# 获取关键列
confirm_time = df.iloc[:, 14]  # 确认接入时间
total_revenue = df.iloc[:, 21] # 总收益
revenue_25 = df.iloc[:, 19]    # 25年总收益
revenue_26 = df.iloc[:, 20]    # 26年总收益

# 当前日期
current_date = datetime(2026, 4, 2)
one_year_ago = datetime(2025, 4, 2)
year_2025_start = datetime(2025, 1, 1)

# 清理数据
confirm_time_clean = pd.to_datetime(confirm_time, errors='coerce')
total_revenue_clean = pd.to_numeric(total_revenue, errors='coerce')
revenue_25_clean = pd.to_numeric(revenue_25, errors='coerce')
revenue_26_clean = pd.to_numeric(revenue_26, errors='coerce')

print('=== 按照新指标定义统计 ===')
print()

# 指标一：接入超过1年，总收益为0
condition1 = (confirm_time_clean < one_year_ago) & ((total_revenue_clean == 0) | (total_revenue_clean.isna()))
zhibiao1 = df[condition1]
print(f'指标一：接入超过1年，总收益为0')
print(f'  数量：{len(zhibiao1)}个')
print()

# 指标二：接入超过1年，0<总收益<10w
condition2 = (confirm_time_clean < one_year_ago) & (total_revenue_clean > 0) & (total_revenue_clean < 100000)
zhibiao2 = df[condition2]
print(f'指标二：接入超过1年，0<总收益<10w')
print(f'  数量：{len(zhibiao2)}个')
print()

# 指标三：25年前接入，0<25年收益<10w
condition3 = (confirm_time_clean < year_2025_start) & (revenue_25_clean > 0) & (revenue_25_clean < 100000)
zhibiao3 = df[condition3]
print(f'指标三：25年前接入，0<25年收益<10w')
print(f'  数量：{len(zhibiao3)}个')
print()

# 指标四：接入超过1年，0<总收益<5w
condition4 = (confirm_time_clean < one_year_ago) & (total_revenue_clean > 0) & (total_revenue_clean < 50000)
zhibiao4 = df[condition4]
print(f'指标四：接入超过1年，0<总收益<5w')
print(f'  数量：{len(zhibiao4)}个')
print()

# 指标五：25年前接入，0<25年收益<5w
condition5 = (confirm_time_clean < year_2025_start) & (revenue_25_clean > 0) & (revenue_25_clean < 50000)
zhibiao5 = df[condition5]
print(f'指标五：25年前接入，0<25年收益<5w')
print(f'  数量：{len(zhibiao5)}个')
print()

# 指标六：26年产生收益
condition6 = (revenue_26_clean > 0) & (revenue_26_clean.notna())
zhibiao6 = df[condition6]
print(f'指标六：26年产生收益')
print(f'  数量：{len(zhibiao6)}个')
print()

# 指标七：25年有收益，26年无
condition7 = (revenue_25_clean > 0) & (revenue_25_clean.notna()) & ((revenue_26_clean == 0) | (revenue_26_clean.isna()))
zhibiao7 = df[condition7]
print(f'指标七：25年有收益，26年无')
print(f'  数量：{len(zhibiao7)}个')
print()

# 分析包含关系
print('=== 包含关系分析 ===')
print()

# 用专区号作为唯一标识
zhibiao1_set = set(zhibiao1.iloc[:, 1].dropna())
zhibiao2_set = set(zhibiao2.iloc[:, 1].dropna())
zhibiao3_set = set(zhibiao3.iloc[:, 1].dropna())
zhibiao4_set = set(zhibiao4.iloc[:, 1].dropna())
zhibiao5_set = set(zhibiao5.iloc[:, 1].dropna())
zhibiao6_set = set(zhibiao6.iloc[:, 1].dropna())
zhibiao7_set = set(zhibiao7.iloc[:, 1].dropna())

# 检查指标四和指标二的关系（指标四应该是指标二的子集）
print(f'指标四是否属于指标二的子集: {zhibiao4_set.issubset(zhibiao2_set)}')
print(f'指标四在指标二中的数量: {len(zhibiao4_set)} / {len(zhibiao2_set)}')
print()

# 检查指标五和指标三的关系（指标五应该是指标三的子集）
print(f'指标五是否属于指标三的子集: {zhibiao5_set.issubset(zhibiao3_set)}')
print(f'指标五在指标三中的数量: {len(zhibiao5_set)} / {len(zhibiao3_set)}')
print()

# 计算独立数量
zhibiao2_only = zhibiao2_set - zhibiao4_set  # 指标二独有的（5-10w）
zhibiao3_only = zhibiao3_set - zhibiao5_set  # 指标三独有的（5-10w）

print('=== 独立数量 ===')
print(f'指标二独有的（5w<=总收益<10w）: {len(zhibiao2_only)}个')
print(f'指标四（0<总收益<5w）: {len(zhibiao4_set)}个')
print(f'验证: {len(zhibiao2_only)} + {len(zhibiao4_set)} = {len(zhibiao2_only) + len(zhibiao4_set)} (应为{len(zhibiao2_set)})')
print()
print(f'指标三独有的（5w<=25年收益<10w）: {len(zhibiao3_only)}个')
print(f'指标五（0<25年收益<5w）: {len(zhibiao5_set)}个')
print(f'验证: {len(zhibiao3_only)} + {len(zhibiao5_set)} = {len(zhibiao3_only) + len(zhibiao5_set)} (应为{len(zhibiao3_set)})')
print()

# 去重后的风险等级统计
print('=== 去重后的风险等级统计 ===')
print()

# 红色：接入超过1年，总收益为0（指标一）
red_zones = zhibiao1_set
print(f'红色-零收益（指标一）: {len(red_zones)}个')

# 橙色：接入超过1年，0<总收益<5w（指标四）
orange_zones = zhibiao4_set - red_zones
print(f'橙色-低收益（指标四）: {len(orange_zones)}个')

# 黄色：接入超过1年，5w<=总收益<10w（指标二独有）
yellow_zones = zhibiao2_only - red_zones - orange_zones
print(f'黄色-中低收益（指标二独有）: {len(yellow_zones)}个')

# 绿色：25年前接入，0<25年收益<5w（指标五）
green_zones = zhibiao5_set - red_zones - orange_zones - yellow_zones
print(f'绿色-25年低收益（指标五）: {len(green_zones)}个')

# 蓝色：25年前接入，5w<=25年收益<10w（指标三独有）
blue_zones = zhibiao3_only - red_zones - orange_zones - yellow_zones - green_zones
print(f'蓝色-25年中收益（指标三独有）: {len(blue_zones)}个')

# 紫色：26年产生收益（指标六）
purple_zones = zhibiao6_set - red_zones - orange_zones - yellow_zones - green_zones - blue_zones
print(f'紫色-26年有收益（指标六）: {len(purple_zones)}个')

# 灰色：25年有收益，26年无（指标七）
grey_zones = zhibiao7_set - red_zones - orange_zones - yellow_zones - green_zones - blue_zones - purple_zones
print(f'灰色-流失风险（指标七）: {len(grey_zones)}个')

# 不重复的总数
all_zones = red_zones | orange_zones | yellow_zones | green_zones | blue_zones | purple_zones | grey_zones
print(f'\n不重复专区总数: {len(all_zones)}个')

# 保存到Excel
print('\n=== 保存分析结果 ===')
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# 样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 工作表1：七大指标统计
ws1 = wb.create_sheet('七大指标统计')
data1 = [
    ['指标', '条件', '数量', '包含关系', '备注'],
    ['指标一', '接入超过1年，总收益为0', len(zhibiao1_set), '-', '新增'],
    ['指标二', '接入超过1年，0<总收益<10w', len(zhibiao2_set), '包含指标四', ''],
    ['指标三', '25年前接入，0<25年收益<10w', len(zhibiao3_set), '包含指标五', ''],
    ['指标四', '接入超过1年，0<总收益<5w', len(zhibiao4_set), '属于指标二', '子集'],
    ['指标五', '25年前接入，0<25年收益<5w', len(zhibiao5_set), '属于指标三', '子集'],
    ['指标六', '26年产生收益', len(zhibiao6_set), '-', ''],
    ['指标七', '25年有收益，26年无', len(zhibiao7_set), '-', ''],
]
for r_idx, row in enumerate(data1, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 15
ws1.column_dimensions['E'].width = 12

# 工作表2：去重后风险等级
ws2 = wb.create_sheet('去重后风险等级')
data2 = [
    ['风险等级', '条件', '数量', '对应指标'],
    ['🔴 红色-零收益', '接入超过1年，总收益为0', len(red_zones), '指标一'],
    ['🟠 橙色-低收益', '接入超过1年，0<总收益<5w', len(orange_zones), '指标四'],
    ['🟡 黄色-中低收益', '接入超过1年，5w<=总收益<10w', len(yellow_zones), '指标二独有'],
    ['🟢 绿色-25年低收益', '25年前接入，0<25年收益<5w', len(green_zones), '指标五'],
    ['🔵 蓝色-25年中收益', '25年前接入，5w<=25年收益<10w', len(blue_zones), '指标三独有'],
    ['🟣 紫色-26年有收益', '26年产生收益', len(purple_zones), '指标六'],
    ['⚪ 灰色-流失风险', '25年有收益，26年无', len(grey_zones), '指标七'],
    ['合计', '不重复专区总数', len(all_zones), ''],
]
for r_idx, row in enumerate(data2, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 15

# 保存
output_path = r'D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\temp\数据分析结果_新指标定义.xlsx'
wb.save(output_path)
print(f'分析结果已保存: {output_path}')
