import openpyxl

# 加载Excel文件
wb = openpyxl.load_workbook('D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\e交易数据监控指标V2.0.xlsx')
sheet = wb.active

print("=" * 80)
print("e交易数据监控指标V2.0 - 详细分析")
print("=" * 80)

# 读取所有数据
data = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
    data.append(row)

# 分析成本监控指标（序号12-17）
print("\n【成本监控指标分析】")
print("-" * 80)

for i in range(11, 17):  # 第12-17行（索引11-16）
    row = data[i]
    print(f"\n指标 {row[0]}: {row[1]} - {row[2]}")
    print(f"  监控规则: {row[3]}")
    print(f"  预警等级: {row[4]}")
    print(f"  统计口径: {row[5]}")

# 分析利润监控指标（序号8-11）
print("\n\n【利润监控指标分析】")
print("-" * 80)

for i in range(7, 11):  # 第8-11行（索引7-10）
    row = data[i]
    print(f"\n指标 {row[0]}: {row[1]} - {row[2]}")
    print(f"  监控规则: {row[3]}")
    print(f"  预警等级: {row[4]}")
    print(f"  统计口径: {row[5]}")

print("\n" + "=" * 80)
