import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook(r'D:\work\龙虾可操作需求临时文件夹\华益评审\企业数字采购产品报价方案新.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== Sheet: {sheet_name} ===')
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = [str(v) if v is not None else '' for v in row]
        print(f'Row {i+1}: {" | ".join(vals)}')
    print()
