import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# 读取CSV数据
df = pd.read_csv('top10_revenue.csv', encoding='utf-8-sig')

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = '25年收益TOP10'

# 设置标题
ws['A1'] = '2025年收益TOP10专区统计表'
ws.merge_cells('A1:F1')
ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# 表头样式
header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
header_font = Font(name='微软雅黑', size=11, bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')

# 边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 写入表头
headers = ['排名', '合同编号', '原专区名称', '所属省份', '总成本', '总收益情况']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 写入数据
data_alignment = Alignment(horizontal='center', vertical='center')
for row_idx, row_data in enumerate(df.itertuples(index=False), start=1):
    # 排名
    ws.cell(row=row_idx+3, column=1, value=row_idx).alignment = data_alignment
    ws.cell(row=row_idx+3, column=1).border = thin_border
    
    # 合同编号
    ws.cell(row=row_idx+3, column=2, value=row_data[0]).alignment = data_alignment
    ws.cell(row=row_idx+3, column=2).border = thin_border
    
    # 原专区名称
    ws.cell(row=row_idx+3, column=3, value=row_data[1]).alignment = data_alignment
    ws.cell(row=row_idx+3, column=3).border = thin_border
    
    # 所属省份
    ws.cell(row=row_idx+3, column=4, value=row_data[2]).alignment = data_alignment
    ws.cell(row=row_idx+3, column=4).border = thin_border
    
    # 总成本
    cost_cell = ws.cell(row=row_idx+3, column=5, value=row_data[3])
    cost_cell.number_format = '#,##0.00'
    cost_cell.alignment = data_alignment
    cost_cell.border = thin_border
    
    # 总收益情况
    revenue_cell = ws.cell(row=row_idx+3, column=6, value=row_data[4])
    revenue_cell.number_format = '#,##0.00'
    revenue_cell.alignment = data_alignment
    revenue_cell.border = thin_border
    
    # 交替行背景色
    if row_idx % 2 == 0:
        for col in range(1, 7):
            ws.cell(row=row_idx+3, column=col).fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

# 设置列宽
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 28
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 18

# 设置行高
for row in range(3, 14):
    ws.row_dimensions[row].height = 22

# 保存文件
output_file = '25年收益TOP10.xlsx'
wb.save(output_file)
print(f'Excel文件已生成: {output_file}')
