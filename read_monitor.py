import openpyxl
import pandas as pd

# 加载Excel文件
wb = openpyxl.load_workbook('D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\e交易数据监控指标V2.0.xlsx')

print("=" * 60)
print("e交易数据监控指标V2.0 - 文件结构分析")
print("=" * 60)

for sheet_name in wb.sheetnames:
    print(f"\nSheet: {sheet_name}")
    sheet = wb[sheet_name]
    
    # 获取数据范围
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"   行数: {max_row}, 列数: {max_col}")
    
    # 读取前30行数据
    data = []
    for row in sheet.iter_rows(min_row=1, max_row=min(30, max_row), values_only=True):
        data.append(row)
    
    # 打印数据
    for i, row in enumerate(data):
        print(f"   行{i+1}: {row}")
    
    if max_row > 30:
        print(f"   ... (还有 {max_row-30} 行数据)")

print("\n" + "=" * 60)
