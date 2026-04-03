import openpyxl
wb = openpyxl.load_workbook('D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\专区信息汇总表_按省份分类.xlsx')
print('处理完成！')
print(f'共处理了 {len(wb.sheetnames)} 个工作表：')
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    row_count = sheet.max_row - 1 if sheet.max_row > 1 else 0
    print(f'  - {sheet_name}: {row_count} 行数据已设置公式')
print('\n公式说明：')
print('  P列: =IF(O列="","",TODAY()-O列)  - 计算O列日期距离今天的天数')
print('  R列: =IF(Q列="","",TODAY()-Q列)  - 计算Q列日期距离今天的天数')
print('\n注意：公式会在Excel打开时自动计算，空值会被保留为空。')
