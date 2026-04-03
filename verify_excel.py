import openpyxl
wb = openpyxl.load_workbook('D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\专区信息汇总表_按省份分类.xlsx')
print('工作表列表:')
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f'  - {sheet_name}: {sheet.max_row}行')
    # 显示P2和R2的公式作为示例
    if sheet.max_row >= 2:
        p2_val = sheet['P2'].value
        r2_val = sheet['R2'].value
        print(f'    P2公式: {p2_val}')
        print(f'    R2公式: {r2_val}')
        break
