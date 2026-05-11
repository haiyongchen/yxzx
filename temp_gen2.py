import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

src = openpyxl.load_workbook(r'D:\work\龙虾可操作需求临时文件夹\华益评审\企业数字采购产品报价方案新.xlsx', data_only=True)
ws_src = src['新点e交易（企业Saas平台）']

selected_items = []
for row in ws_src.iter_rows(min_row=2, max_row=120, values_only=True):
    if row and len(row) >= 10:
        seq = row[0]
        suite = str(row[1]).strip() if row[1] else ''
        required = str(row[2]).strip() if row[2] else ''
        module = str(row[3]).strip() if row[3] else ''
        selected = str(row[5]).strip() if row[5] else ''
        unit_price = row[6] if row[6] else 0
        qty = row[7] if row[7] else 0
        saas_price = row[8] if row[8] else 0
        deploy_price = row[9] if row[9] else 0
        desc = str(row[10]).strip() if row[10] else ''
        if selected == '√' and seq and str(seq).isdigit():
            selected_items.append({
                'seq': int(seq), 'suite': suite, 'required': required,
                'module': module,
                'unit_price': unit_price if isinstance(unit_price, (int, float)) else 0,
                'qty': qty if isinstance(qty, (int, float)) else 0,
                'saas': saas_price if isinstance(saas_price, (int, float)) else 0,
                'deploy': deploy_price if isinstance(deploy_price, (int, float)) else 0,
                'desc': desc
            })

wb = openpyxl.Workbook()
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14)
normal_font = Font(name='微软雅黑', size=10)
required_font = Font(name='微软雅黑', size=10, color='C00000', bold=True)
optional_font = Font(name='微软雅黑', size=10, color='2F5496')
money_font = Font(name='微软雅黑', size=10, bold=True, color='C00000')
total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
opt_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)

def write_sheet(ws, title, price_col_name, price_key):
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A2:H2')
    ws['A2'] = '报价单位：新点软件 | 含税（6%） | 价格单位：元'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['序号', '功能套件', '功能模块', '功能说明', '选型', '单价', '数量', price_col_name]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row = 5
    for item in selected_items:
        vals = [item['seq'], item['suite'], item['module'],
                item['desc'][:50] if item['desc'] else '',
                item['required'], item['unit_price'], item['qty'], item[price_key]]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align if col in [1, 5, 6, 7, 8] else left_align
            if col == 5:
                cell.font = required_font if val == '必选' else optional_font
            elif col == 8 and val and isinstance(val, (int, float)) and val > 0:
                cell.font = money_font
                cell.number_format = '#,##0'
            elif col == 6 and val and isinstance(val, (int, float)) and val > 0:
                cell.font = normal_font
                cell.number_format = '#,##0'
            else:
                cell.font = normal_font
        row += 1

    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1, value='必选功能小计（含税6%）').font = Font(name='微软雅黑', bold=True, size=11, color='C00000')
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).alignment = right_align
    req_sum = sum(item[price_key] for item in selected_items if item['required'] == '必选')
    ws.cell(row=row, column=8, value=req_sum).font = Font(name='微软雅黑', bold=True, size=11, color='C00000')
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).number_format = '#,##0'
    ws.cell(row=row, column=8).alignment = center_align
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = thin_border

    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1, value='可选功能小计（含税6%）').font = Font(name='微软雅黑', bold=True, size=11, color='2F5496')
    ws.cell(row=row, column=1).fill = opt_fill
    ws.cell(row=row, column=1).alignment = right_align
    opt_sum = sum(item[price_key] for item in selected_items if item['required'] != '必选')
    ws.cell(row=row, column=8, value=opt_sum).font = Font(name='微软雅黑', bold=True, size=11, color='2F5496')
    ws.cell(row=row, column=8).fill = opt_fill
    ws.cell(row=row, column=8).number_format = '#,##0'
    ws.cell(row=row, column=8).alignment = center_align
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = thin_border

    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    ws.cell(row=row, column=1, value='报价合计（含税6%）').font = Font(name='微软雅黑', bold=True, size=14, color='C00000')
    ws.cell(row=row, column=1).fill = total_fill
    ws.cell(row=row, column=1).alignment = right_align
    total = req_sum + opt_sum
    ws.cell(row=row, column=8, value=total).font = Font(name='微软雅黑', bold=True, size=14, color='C00000')
    ws.cell(row=row, column=8).fill = total_fill
    ws.cell(row=row, column=8).number_format = '#,##0'
    ws.cell(row=row, column=8).alignment = center_align
    for c in range(1, 9):
        ws.cell(row=row, column=c).border = thin_border

    row += 2
    ws.cell(row=row, column=1, value='备注：').font = Font(name='微软雅黑', bold=True, size=10)
    row += 1
    for note in ['1. 红色"必选"为必选模块，蓝色"可选"为可选模块',
                 '2. 以上价格为标准报价，实际可根据项目情况调整',
                 '3. 用户数量暂不按此收费，不限用户数',
                 '4. 交付服务（实施/培训/定制开发）费用按实际情况另行报价']:
        ws.cell(row=row, column=1, value=note).font = Font(name='微软雅黑', size=9, color='666666')
        ws.merge_cells(f'A{row}:H{row}')
        row += 1

    col_widths = [6, 18, 28, 40, 8, 12, 8, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

ws1 = wb.active
ws1.title = 'SaaS版报价（年度订阅）'
write_sheet(ws1, '华益电子招采平台 — SaaS版报价清单（年度订阅）', 'SaaS报价/年', 'saas')

ws2 = wb.create_sheet('私有化部署版报价')
write_sheet(ws2, '华益电子招采平台 — 私有化部署版报价清单', '私有化部署', 'deploy')

# Save to temp first then copy
import shutil
temp_path = r'D:\openclaw-workspace\temp_quotation.xlsx'
wb.save(temp_path)
target_path = r'D:\work\龙虾可操作需求临时文件夹\华益评审\华益电子招采平台报价清单_两版方案.xlsx'
try:
    shutil.copy2(temp_path, target_path)
    print(f'文件已生成: {target_path}')
except:
    alt_path = r'D:\work\龙虾可操作需求临时文件夹\华益评审\华益电子招采平台报价清单_两版方案_新.xlsx'
    shutil.copy2(temp_path, alt_path)
    print(f'文件已生成: {alt_path}')

# Summary
print(f'\n共选中 {len(selected_items)} 项')
saas_req = sum(i['saas'] for i in selected_items if i['required'] == '必选')
saas_opt = sum(i['saas'] for i in selected_items if i['required'] != '必选')
deploy_req = sum(i['deploy'] for i in selected_items if i['required'] == '必选')
deploy_opt = sum(i['deploy'] for i in selected_items if i['required'] != '必选')
print(f'SaaS版 — 必选: {saas_req:,} | 可选: {saas_opt:,} | 合计: {saas_req+saas_opt:,}')
print(f'部署版 — 必选: {deploy_req:,} | 可选: {deploy_opt:,} | 合计: {deploy_req+deploy_opt:,}')
