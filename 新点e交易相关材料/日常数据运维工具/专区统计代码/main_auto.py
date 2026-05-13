import openpyxl
import requests
import os
import datetime


def process_excel_auto(log_func=print):
    """
    自动处理固定路径的Excel文件
    输入文件: D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\同步数据文件\专区接入情况统计表.xlsx
    输出模式: 原地覆盖
    """
    # 固定输入文件路径
    in_path = r"D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\同步数据文件\专区接入情况统计表.xlsx"
    api_path = "json/ZiZhanIndexCount.json"
    
    # 检查文件是否存在
    if not os.path.exists(in_path):
        log_func(f"[错误] 文件不存在: {in_path}")
        return None
    
    try:
        log_func(f"正在读取: {os.path.basename(in_path)}")
        wb = openpyxl.load_workbook(in_path)
        
        # 原地覆盖模式
        save_path = in_path
        log_func(f"保存模式：原地覆盖")

        for sheet in wb.worksheets:
            if sheet.max_row < 2: 
                continue
            log_func(f"--- 扫描分表: {sheet.title} ---")

            # --- 1. 寻找现有列 ---
            addr_col, jy_col, kb_col = None, None, None
            last_filled_col = 0

            # 遍历当前已有的所有列，寻找表头
            for col in range(1, sheet.max_column + 1):
                cell_val = sheet.cell(row=1, column=col).value
                if cell_val:
                    last_filled_col = col
                    header_str = str(cell_val).replace(" ", "")
                    if "专区地址" in header_str: 
                        addr_col = col
                    if "近期交易" in header_str: 
                        jy_col = col
                    if "今日开标" in header_str: 
                        kb_col = col

            if not addr_col:
                log_func("跳过：未找到'专区地址'列")
                continue

            # --- 2. 确定写入列的位置 ---
            if not jy_col:
                jy_col = last_filled_col + 1
                kb_col = last_filled_col + 2
                sheet.cell(row=1, column=jy_col).value = "近期交易"
                sheet.cell(row=1, column=kb_col).value = "今日开标"
                log_func(f"在数据边缘新建列: 第 {jy_col} 和 {kb_col} 列")
            else:
                if not kb_col: 
                    kb_col = jy_col + 1
                log_func(f"匹配并覆盖现有列: 第 {jy_col} 和 {kb_col} 列")

            # --- 3. 执行 API 请求与数据回填 ---
            success_count = 0
            error_count = 0
            
            for row in range(2, sheet.max_row + 1):
                url_val = sheet.cell(row=row, column=addr_col).value
                if not url_val or "http" not in str(url_val): 
                    continue

                full_url = str(url_val).strip()
                if not full_url.endswith('/'): 
                    full_url += '/'

                try:
                    resp = requests.get(f"{full_url}{api_path}", timeout=5)
                    if resp.status_code == 200:
                        data = resp.json()
                        sheet.cell(row=row, column=jy_col).value = data.get("countjy", 0)
                        sheet.cell(row=row, column=kb_col).value = data.get("countkb", 0)
                        success_count += 1
                    else:
                        sheet.cell(row=row, column=jy_col).value = f"HTTP:{resp.status_code}"
                        error_count += 1
                except Exception as e:
                    sheet.cell(row=row, column=jy_col).value = "超时或错误"
                    error_count += 1

            log_func(f"分表 [{sheet.title}] 更新完成 (成功: {success_count}条, 失败: {error_count}条)")

        # --- 4. 保存 (原地覆盖) ---
        try:
            wb.save(save_path)
            log_func(f"\n[成功] 任务执行完毕！")
            log_func(f"文件已保存: {save_path}")
            return save_path
        except PermissionError:
            log_func("\n[错误] 保存失败！请确保 Excel 文件已关闭。")
            return None
            
    except Exception as e:
        log_func(f"运行出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


if __name__ == "__main__":
    print("=" * 60)
    print("BiddingCount 自动处理程序")
    print("=" * 60)
    print(f"输入文件: D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\同步数据文件\\专区接入情况统计表.xlsx")
    print("输出模式: 原地覆盖")
    print("=" * 60)
    print()
    
    result = process_excel_auto()
    
    if result:
        print(f"\n✓ 处理成功: {result}")
    else:
        print("\n✗ 处理失败")
