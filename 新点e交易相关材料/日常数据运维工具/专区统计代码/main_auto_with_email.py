#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BiddingCount 自动处理程序 (带邮件通知)
处理多个Excel文件并发送邮件通知
"""

import openpyxl
import requests
import os
import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


def send_email(subject, body):
    """发送邮件通知"""
    smtp_server = "smtp.qq.com"
    smtp_port = 465
    sender_email = "631115784@qq.com"
    sender_password = "jmxrfcsmhxzabfec"
    receiver_email = "631115784@qq.com"
    
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = f"[专区数据同步任务执行] {subject}"
    
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    
    try:
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, receiver_email, msg.as_string())
        return True
    except Exception as e:
        print(f"邮件发送失败: {e}")
        return False


def process_single_file(in_path, log_func=print):
    """处理单个Excel文件"""
    api_path = "json/ZiZhanIndexCount.json"
    
    if not os.path.exists(in_path):
        log_func(f"[错误] 文件不存在: {in_path}")
        return None, 0, 0
    
    try:
        log_func(f"正在读取: {os.path.basename(in_path)}")
        wb = openpyxl.load_workbook(in_path)
        save_path = in_path
        
        total_success = 0
        total_error = 0
        sheet_results = []

        for sheet in wb.worksheets:
            if sheet.title == "落地":
                log_func(f"--- 跳过sheet: {sheet.title} ---")
                continue
            
            if sheet.max_row < 2: 
                continue
            log_func(f"--- 扫描sheet: {sheet.title} ---")

            # 寻找列
            addr_col, jy_col, kb_col = None, None, None
            last_filled_col = 0

            for col in range(1, sheet.max_column + 1):
                cell_val = sheet.cell(row=1, column=col).value
                if cell_val:
                    last_filled_col = col
                    header_str = str(cell_val).replace(" ", "")
                    if "专区地址" in header_str: 
                        addr_col = col
                    if "近期交易" in header_str: 
                        jy_col = col
                    if "今日开标" in header_str: 
                        kb_col = col

            if not addr_col:
                log_func("跳过：未找到'专区地址'列")
                continue

            # 确定写入列
            if not jy_col:
                jy_col = last_filled_col + 1
                kb_col = last_filled_col + 2
                sheet.cell(row=1, column=jy_col).value = "近期交易"
                sheet.cell(row=1, column=kb_col).value = "今日开标"
                log_func(f"新建列: 第 {jy_col} 和 {kb_col} 列")
            else:
                if not kb_col: 
                    kb_col = jy_col + 1
                log_func(f"覆盖现有列: 第 {jy_col} 和 {kb_col} 列")

            # 执行API请求
            success_count = 0
            error_count = 0
            
            for row in range(2, sheet.max_row + 1):
                url_val = sheet.cell(row=row, column=addr_col).value
                if not url_val or "http" not in str(url_val): 
                    continue

                full_url = str(url_val).strip()
                if not full_url.endswith('/'): 
                    full_url += '/'

                try:
                    resp = requests.get(f"{full_url}{api_path}", timeout=5)
                    if resp.status_code == 200:
                        data = resp.json()
                        sheet.cell(row=row, column=jy_col).value = data.get("countjy", 0)
                        sheet.cell(row=row, column=kb_col).value = data.get("countkb", 0)
                        success_count += 1
                    else:
                        sheet.cell(row=row, column=jy_col).value = f"HTTP:{resp.status_code}"
                        error_count += 1
                except Exception as e:
                    sheet.cell(row=row, column=jy_col).value = "超时"
                    error_count += 1

            log_func(f"Sheet [{sheet.title}] 完成 (成功: {success_count}, 失败: {error_count})")
            sheet_results.append(f"{sheet.title}: 成功{success_count}, 失败{error_count}")
            total_success += success_count
            total_error += error_count

        # 保存
        try:
            wb.save(save_path)
            log_func(f"[成功] 文件已保存: {save_path}")
            return save_path, total_success, total_error, sheet_results
        except PermissionError:
            log_func("[错误] 保存失败！请确保 Excel 文件已关闭。")
            return None, 0, 0, []
            
    except Exception as e:
        log_func(f"运行出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, 0, 0, []


def main():
    start_time = datetime.datetime.now()
    
    print("=" * 60)
    print("BiddingCount 自动处理程序 (带邮件通知)")
    print("=" * 60)
    
    # 定义要处理的文件
    files = [
        r"D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\同步数据文件\专区接入情况统计表.xlsx",
        r"D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\同步数据文件\专区信息汇总表_按省份分类.xlsx"
    ]
    
    print(f"需要处理的文件: {len(files)} 个")
    print("跳过: 落地 sheet")
    print("=" * 60)
    print()
    
    all_results = []
    total_success = 0
    total_error = 0
    
    for i, file_path in enumerate(files, 1):
        print(f"\n{'='*60}")
        print(f"[{i}/{len(files)}] 处理: {os.path.basename(file_path)}")
        print(f"{'='*60}")
        
        result, success, error, sheets = process_single_file(file_path)
        
        if result:
            print(f"[成功] {result}")
            all_results.append({
                'file': os.path.basename(file_path),
                'success': success,
                'error': error,
                'sheets': sheets
            })
            total_success += success
            total_error += error
        else:
            print(f"[失败]")
            all_results.append({
                'file': os.path.basename(file_path),
                'success': 0,
                'error': 0,
                'sheets': []
            })
    
    end_time = datetime.datetime.now()
    duration = (end_time - start_time).total_seconds()
    
    print("\n" + "="*60)
    print("所有文件处理完成!")
    print("="*60)
    
    # 构建邮件内容
    email_body = f"""
专区数据同步任务执行报告

执行时间: {start_time.strftime('%Y-%m-%d %H:%M:%S')}
执行时长: {duration:.1f} 秒
任务状态: 执行完成

处理结果汇总:
"""
    
    for result in all_results:
        email_body += f"\n文件: {result['file']}\n"
        email_body += f"  成功: {result['success']} 条\n"
        email_body += f"  失败: {result['error']} 条\n"
        if result['sheets']:
            email_body += "  Sheet详情:\n"
            for sheet in result['sheets']:
                email_body += f"    - {sheet}\n"
    
    email_body += f"\n总计: 成功 {total_success} 条, 失败 {total_error} 条\n"
    email_body += "\n---\n此邮件由系统自动发送\n"
    
    # 发送邮件
    print("\n正在发送邮件通知...")
    email_sent = send_email(
        subject=f"完成 - 成功{total_success}条",
        body=email_body
    )
    
    if email_sent:
        print("[成功] 邮件已发送")
    else:
        print("[失败] 邮件发送失败")
    
    print("\n" + "="*60)
    print("任务全部完成!")
    print("="*60)


if __name__ == "__main__":
    main()
