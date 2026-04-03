import openpyxl
from datetime import datetime

# 加载Excel文件
wb = openpyxl.load_workbook('D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\专区信息汇总表_按省份分类.xlsx')

# 获取当前日期用于公式计算
today = datetime.now()

# 遍历所有sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f'处理Sheet: {sheet_name}')
    
    # 获取数据行数
    max_row = sheet.max_row
    
    # 从第2行开始设置公式（假设第1行是表头）
    for row in range(2, max_row + 1):
        # P列: 计算O列距离当前时间的天数，兼容空值
        # 公式: =IF(O2="","",TODAY()-O2)
        sheet[f'P{row}'] = f'=IF(O{row}="","",TODAY()-O{row})'
        
        # R列: 计算Q列距离当前时间的天数，兼容空值
        # 公式: =IF(Q2="","",TODAY()-Q2)
        sheet[f'R{row}'] = f'=IF(Q{row}="","",TODAY()-Q{row})'
    
    print(f'  已设置P列和R列公式，共{max_row-1}行数据')

# 保存文件
wb.save('D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\专区信息汇总表_按省份分类.xlsx')
print('\n文件保存成功！')
