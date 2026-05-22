import openpyxl
import sys
from copy import copy

sys.stdout.reconfigure(encoding='utf-8')

# 电商名单
ecommerce_list = [
    '苏宁易购集团股份有限公司',
    '得力集团有限公司',
    '欧菲斯集团股份有限公司',
    '深圳齐心集团股份有限公司',
    '大江科技集团有限公司',
    '史泰博(上海)有限公司',
    '阳采集团有限公司',
    '江苏比高机电设备有限公司',
    '浙江宏伟供应链集团股份有限公司',
    '深圳市怡亚通供应链股份有限公司',
    '咸亨国际科技股份有限公司',
    '紫迈电子商务有限公司',
    '鑫方盛数智科技股份有限公司',
    '震坤行工业超市（上海）有限公司'
]

file_path = r'D:\work\龙虾可操作需求临时文件夹\订单数据处理\阳光优采交易订单.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print('=== 处理前供应商类型分布 ===')
# 统计J列（第10列）的值
type_count = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=10, values_only=True):
    val = row[0]
    if val:
        type_count[val] = type_count.get(val, 0) + 1
print(type_count)
print()

# 创建合并单元格映射：记录每个合并区域
merged_map = {}
for merged_range in ws.merged_cells.ranges:
    min_row = merged_range.min_row
    max_row = merged_range.max_row
    min_col = merged_range.min_col
    max_col = merged_range.max_col
    # 对于合并区域中的每个单元格，记录其应该使用的值（左上角单元格）
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            if row != min_row or col != min_col:
                merged_map[(row, col)] = (min_row, min_col)

print(f'合并单元格映射数量: {len(merged_map)}')
print()

# 先解除所有合并单元格，填充值
print('解除合并单元格并填充值...')
for merged_range in list(ws.merged_cells.ranges):
    min_row = merged_range.min_row
    max_row = merged_range.max_row
    min_col = merged_range.min_col
    max_col = merged_range.max_col
    
    # 获取左上角单元格的值
    top_left_value = ws.cell(row=min_row, column=min_col).value
    
    # 解除合并
    ws.unmerge_cells(str(merged_range))
    
    # 填充所有单元格为左上角的值
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).value = top_left_value

print('合并单元格已解除')
print()

# 现在处理供应商类型（J列，第10列）
# 供应商在I列（第9列）
print('处理供应商类型...')
updated_count = 0
for row in range(2, ws.max_row + 1):
    supplier = ws.cell(row=row, column=9).value  # I列：供应商
    
    if supplier:
        supplier = str(supplier).strip()
        # 判断是否是电商
        is_ecommerce = False
        for ecom in ecommerce_list:
            if supplier == ecom:
                is_ecommerce = True
                break
            # 处理"史泰博"简称
            if '史泰博' in supplier and '史泰博' in ecom:
                is_ecommerce = True
                break
        
        new_type = '电商供应商' if is_ecommerce else '本地供应商'
    else:
        new_type = '本地供应商'
    
    old_type = ws.cell(row=row, column=10).value
    if old_type != new_type:
        ws.cell(row=row, column=10).value = new_type
        updated_count += 1

print(f'更新了 {updated_count} 行供应商类型')
print()

# 统计处理后的分布
print('=== 处理后供应商类型分布 ===')
type_count = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=10, max_col=10, values_only=True):
    val = row[0]
    if val:
        type_count[val] = type_count.get(val, 0) + 1
print(type_count)
print()

# 保存文件
output_path = r'D:\work\龙虾可操作需求临时文件夹\订单数据处理\阳光优采交易订单_处理后.xlsx'
wb.save(output_path)
print(f'文件已保存到: {output_path}')
