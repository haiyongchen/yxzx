import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'D:\work\龙虾可操作需求临时文件夹\订单数据处理\阳光优采交易订单 (38).xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 记录合并单元格
merged_ranges = list(ws.merged_cells.ranges)
# 创建一个集合，记录哪些行是"从属行"（合并单元格中非首行）
slave_rows = set()
for mr in merged_ranges:
    for r in range(mr.min_row + 1, mr.max_row + 1):
        slave_rows.add(r)

print(f'合并单元格数量: {len(merged_ranges)}')
print(f'从属行数量: {len(slave_rows)}')
print()

# 统计
ecom_count = 0
ecom_amount = 0
local_count = 0
local_amount = 0
empty_count = 0
empty_amount = 0

for row_idx in range(2, ws.max_row + 1):
    # 跳过从属行（合并单元格中的非首行）
    if row_idx in slave_rows:
        continue
    
    supplier_type = ws.cell(row=row_idx, column=10).value  # J列：供应商类型
    amount = ws.cell(row=row_idx, column=15).value         # O列：订单金额
    
    # 处理金额
    if amount is None or amount == '':
        amount = 0
    try:
        amount = float(amount)
    except:
        amount = 0
    
    if supplier_type == '电商供应商':
        ecom_count += 1
        ecom_amount += amount
    elif supplier_type == '本地供应商':
        local_count += 1
        local_amount += amount
    else:
        empty_count += 1
        empty_amount += amount

print('=== 按供应商类型统计（合并单元格按1单计算）===')
print(f'电商供应商: {ecom_count} 单, 金额 {ecom_amount:,.2f} 元')
print(f'本地供应商: {local_count} 单, 金额 {local_amount:,.2f} 元')
if empty_count > 0:
    print(f'未填写: {empty_count} 单, 金额 {empty_amount:,.2f} 元')
print()
total = ecom_count + local_count + empty_count
total_amount = ecom_amount + local_amount + empty_amount
print(f'总计: {total} 单, 金额 {total_amount:,.2f} 元')

wb.close()
