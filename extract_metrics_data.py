#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按照新指标定义提取原始数据
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 读取原始数据
file_path = r'D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\temp\专区信息汇总表_中原华北_合并.xlsx'
df = pd.read_excel(file_path)

# 获取关键列
confirm_time = df.iloc[:, 14]  # 确认接入时间
total_revenue = df.iloc[:, 21] # 总收益
revenue_25 = df.iloc[:, 19]    # 25年总收益
revenue_26 = df.iloc[:, 20]    # 26年总收益
total_cost = df.iloc[:, 25]    # 总成本

# 当前日期
current_date = datetime(2026, 4, 2)
one_year_ago = datetime(2025, 4, 2)
year_2025_start = datetime(2025, 1, 1)

# 清理数据
confirm_time_clean = pd.to_datetime(confirm_time, errors='coerce')
total_revenue_clean = pd.to_numeric(total_revenue, errors='coerce')
revenue_25_clean = pd.to_numeric(revenue_25, errors='coerce')
revenue_26_clean = pd.to_numeric(revenue_26, errors='coerce')

# 要导出的列
cols_to_export = [df.columns[0], df.columns[1], df.columns[2], df.columns[14], 
                  df.columns[21], df.columns[19], df.columns[20], df.columns[25]]
header_names = ['合同编号', '专区号', '专区名称', '确认接入时间', 
                '总收益', '25年总收益', '26年总收益', '总成本']

# 指标一：接入超过1年，总收益为0
condition1 = (confirm_time_clean < one_year_ago) & ((total_revenue_clean == 0) | (total_revenue_clean.isna()))
zhibiao1 = df[condition1][cols_to_export].copy()
zhibiao1.columns = header_names

# 指标二：接入超过1年，0<总收益<10w
condition2 = (confirm_time_clean < one_year_ago) & (total_revenue_clean > 0) & (total_revenue_clean < 100000)
zhibiao2 = df[condition2][cols_to_export].copy()
zhibiao2.columns = header_names

# 指标三：25年前接入，0<25年收益<10w
condition3 = (confirm_time_clean < year_2025_start) & (revenue_25_clean > 0) & (revenue_25_clean < 100000)
zhibiao3 = df[condition3][cols_to_export].copy()
zhibiao3.columns = header_names

# 指标四：接入超过1年，0<总收益<5w
condition4 = (confirm_time_clean < one_year_ago) & (total_revenue_clean > 0) & (total_revenue_clean < 50000)
zhibiao4 = df[condition4][cols_to_export].copy()
zhibiao4.columns = header_names

# 指标五：25年前接入，0<25年收益<5w
condition5 = (confirm_time_clean < year_2025_start) & (revenue_25_clean > 0) & (revenue_25_clean < 50000)
zhibiao5 = df[condition5][cols_to_export].copy()
zhibiao5.columns = header_names

# 指标六：26年产生收益
condition6 = (revenue_26_clean > 0) & (revenue_26_clean.notna())
zhibiao6 = df[condition6][cols_to_export].copy()
zhibiao6.columns = header_names

# 指标七：25年有收益，26年无
condition7 = (revenue_25_clean > 0) & (revenue_25_clean.notna()) & ((revenue_26_clean == 0) | (revenue_26_clean.isna()))
zhibiao7 = df[condition7][cols_to_export].copy()
zhibiao7.columns = header_names

# 创建Excel工作簿
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# 样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def write_data_to_sheet(ws, data, sheet_name):
    """写入数据到工作表"""
    # 写入表头
    for c_idx, header in enumerate(header_names, 1):
        cell = ws.cell(row=1, column=c_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # 写入数据
    for r_idx, row in enumerate(data.values, 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if c_idx in [5, 6, 7, 8]:  # 数值列居中
                cell.alignment = Alignment(horizontal='center')
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

# 1. 指标一：接入超过1年，总收益为0
ws1 = wb.create_sheet('指标一-零收益')
write_data_to_sheet(ws1, zhibiao1, '指标一')

# 2. 指标二：接入超过1年，0<总收益<10w
ws2 = wb.create_sheet('指标二-总收益0-10w')
write_data_to_sheet(ws2, zhibiao2, '指标二')

# 3. 指标三：25年前接入，0<25年收益<10w
ws3 = wb.create_sheet('指标三-25年收益0-10w')
write_data_to_sheet(ws3, zhibiao3, '指标三')

# 4. 指标四：接入超过1年，0<总收益<5w
ws4 = wb.create_sheet('指标四-总收益0-5w')
write_data_to_sheet(ws4, zhibiao4, '指标四')

# 5. 指标五：25年前接入，0<25年收益<5w
ws5 = wb.create_sheet('指标五-25年收益0-5w')
write_data_to_sheet(ws5, zhibiao5, '指标五')

# 6. 指标六：26年产生收益
ws6 = wb.create_sheet('指标六-26年有收益')
write_data_to_sheet(ws6, zhibiao6, '指标六')

# 7. 指标七：25年有收益，26年无
ws7 = wb.create_sheet('指标七-25有26无')
write_data_to_sheet(ws7, zhibiao7, '指标七')

# 8. 汇总统计（放在第一个）
ws8 = wb.create_sheet('汇总统计', 0)
summary_data = [
    ['指标', '条件', '数量', '说明'],
    ['指标一', '接入超过1年，总收益为0', len(zhibiao1), '新增'],
    ['指标二', '接入超过1年，0<总收益<10w', len(zhibiao2), '包含指标四'],
    ['指标三', '25年前接入，0<25年收益<10w', len(zhibiao3), '包含指标五'],
    ['指标四', '接入超过1年，0<总收益<5w', len(zhibiao4), '属于指标二'],
    ['指标五', '25年前接入，0<25年收益<5w', len(zhibiao5), '属于指标三'],
    ['指标六', '26年产生收益', len(zhibiao6), ''],
    ['指标七', '25年有收益，26年无', len(zhibiao7), ''],
]
for r_idx, row in enumerate(summary_data, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws8.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
ws8.column_dimensions['A'].width = 12
ws8.column_dimensions['B'].width = 40
ws8.column_dimensions['C'].width = 10
ws8.column_dimensions['D'].width = 15

# 保存
output_path = r'D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\temp\七大指标原始数据.xlsx'
wb.save(output_path)

print('=== 七大指标数据提取完成 ===')
print(f'文件路径: {output_path}')
print()
print('包含工作表:')
print(f'1. 汇总统计 - 7个指标的数量汇总')
print(f'2. 指标一-零收益 - {len(zhibiao1)}条')
print(f'3. 指标二-总收益0-10w - {len(zhibiao2)}条')
print(f'4. 指标三-25年收益0-10w - {len(zhibiao3)}条')
print(f'5. 指标四-总收益0-5w - {len(zhibiao4)}条')
print(f'6. 指标五-25年收益0-5w - {len(zhibiao5)}条')
print(f'7. 指标六-26年有收益 - {len(zhibiao6)}条')
print(f'8. 指标七-25有26无 - {len(zhibiao7)}条')
print()
print('每个工作表包含字段:')
print('  - 合同编号、专区号、专区名称、确认接入时间')
print('  - 总收益、25年总收益、26年总收益、总成本')
