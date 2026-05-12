import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
import os

src = r'D:\work\龙虾可操作需求临时文件夹\企业采购监管数据质检模型.xlsx'
out_dir = r'D:\work\龙虾可操作需求临时文件夹'

wb = openpyxl.load_workbook(src)

# Create two new workbooks
wb_l1 = openpyxl.Workbook()
wb_l2 = openpyxl.Workbook()

# Remove default sheets
wb_l1.remove(wb_l1.active)
wb_l2.remove(wb_l2.active)

# Define styles
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

table_names = ['表1', '表2', '表3', '表4', '表5', '表6', '表7', '表8']

for tname in table_names:
    ws = wb[tname]
    
    # Collect rows by level
    l1_rows = []
    l2_rows = []
    
    for r in range(4, ws.max_row + 1):
        level = ws.cell(row=r, column=5).value  # E列 = 校验等级
        if level is None:
            continue
        level_str = str(level).strip()
        
        row_data = []
        for c in range(1, 11):  # A到J列
            row_data.append(ws.cell(row=r, column=c).value)
        
        if level_str == '一级':
            l1_rows.append(row_data)
        elif level_str == '二级':
            l2_rows.append(row_data)
    
    # Write to Level 1 workbook
    ws1 = wb_l1.create_sheet(title=tname)
    # Copy header rows (1-3)
    for r in range(1, 4):
        for c in range(1, 11):
            src_cell = ws.cell(row=r, column=c)
            dst_cell = ws1.cell(row=r, column=c, value=src_cell.value)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.border = copy(src_cell.border)
    # Write data rows
    for i, row_data in enumerate(l1_rows):
        for c, val in enumerate(row_data):
            cell = ws1.cell(row=4+i, column=c+1, value=val)
            cell.border = thin_border
    
    # Write to Level 2 workbook
    ws2 = wb_l2.create_sheet(title=tname)
    for r in range(1, 4):
        for c in range(1, 11):
            src_cell = ws.cell(row=r, column=c)
            dst_cell = ws2.cell(row=r, column=c, value=src_cell.value)
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font)
                dst_cell.fill = copy(src_cell.fill)
                dst_cell.alignment = copy(src_cell.alignment)
                dst_cell.border = copy(src_cell.border)
    for i, row_data in enumerate(l2_rows):
        for c, val in enumerate(row_data):
            cell = ws2.cell(row=4+i, column=c+1, value=val)
            cell.border = thin_border

# Save
out1 = os.path.join(out_dir, '企业采购监管数据质检模型_一级监管.xlsx')
out2 = os.path.join(out_dir, '企业采购监管数据质检模型_二级监管.xlsx')

wb_l1.save(out1)
wb_l2.save(out2)

print(f'拆分完成！')
print(f'一级监管: {out1}')
print(f'二级监管: {out2}')

# Summary
for tname in table_names:
    ws = wb[tname]
    l1_count = 0
    l2_count = 0
    for r in range(4, ws.max_row + 1):
        level = ws.cell(row=r, column=5).value
        if level is None:
            continue
        level_str = str(level).strip()
        if level_str == '一级':
            l1_count += 1
        elif level_str == '二级':
            l2_count += 1
    print(f'{tname}: 一级 {l1_count} 条, 二级 {l2_count} 条')
