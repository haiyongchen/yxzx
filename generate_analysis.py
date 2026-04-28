from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# Sheet1: 0分原因深度分析
ws1 = wb.active
ws1.title = '0分原因深度分析'

headers1 = ['评分点', '0分原因分类', '具体表现', '问题本质', '1分标准要点', 'AI识别关键词', '规则权重']
ws1.append(headers1)

data_analysis = [
    ['综合阐述', '目标不清晰', '改造目标描述模糊', '缺乏明确的项目目标定位', '目标明确、文字清楚', '目标、明确、清晰', '高'],
    ['综合阐述', '内容偏离', '改造内容未写安装保温材料', '内容与实际需求不符', '内容完整、符合要求', '保温材料、内容缺失', '高'],
    ['综合阐述', '表述冗长晦涩', '文字内容晦涩冗长', '可读性差，表达不清晰', '文字清楚、简洁明了', '晦涩、冗长、不清楚', '中'],
    ['综合阐述', '内容混乱', '包含无关内容（融资、股权抵押、气瓶储存等）', '内容混杂，缺乏针对性', '针对本项目、内容相关', '融资、股权、抵押、气瓶、危险化学品', '高'],
    ['综合阐述', '针对性不足', '针对本标段的概述欠缺/不详细', '缺乏对本项目的具体分析', '针对本标段、具体阐述', '欠缺、不详细、针对性', '高'],
    ['综合阐述', '范围错误', '综合阐述范围不是本标段改造范围', '项目范围理解错误', '范围正确、符合标段', '范围错误、不是本标段', '高'],
    ['综合阐述', '信息错误', '针对本项目的名称及范围有误', '基本信息错误', '名称正确、范围准确', '有误、错误', '高'],
    ['综合阐述', '完全缺失', '无针对本小区的具体内容', '完全没有针对性内容', '有针对性内容', '无、没有、缺乏', '高'],
    ['管理架构', '图表错误', '组织架构图错误', '图表本身存在错误', '架构正确、图表清晰', '错误、不正确', '高'],
    ['管理架构', '描述不具体', '对人员的岗位职责描述不够具体', '职责描述空泛', '职责明确、描述具体', '不够具体、不详细', '中'],
    ['管理架构', '内容不全面', '内容不全面，过于简单', '内容缺失，过于简略', '内容全面、详细', '不全面、过于简单', '中'],
    ['管理架构', '针对性不足', '针对本标段的安排欠缺/不详细/不明确', '缺乏对本项目的针对性', '针对本标段、具体安排', '欠缺、不详细、不明确', '高'],
    ['施工进度计划', '描述不详细', '对横道图详细描述不够/不细致', '图表说明不充分', '描述详细、清晰', '不够、不细致', '中'],
    ['施工进度计划', '计划性说明不足', '未能对横道图有计划性说明', '缺乏对计划的详细阐述', '有计划性说明', '未能、没有说明', '高'],
    ['施工进度计划', '阶段划分不清', '分为六个阶段并未能标注日历天数', '时间节点不明确', '阶段清晰、时间明确', '未能标注、天数', '高'],
    ['施工进度计划', '表述不清晰', '文字表述计划不清晰明确', '文字表达问题', '表述清晰明确', '不清晰、不明确', '中'],
    ['施工进度计划', '工期目标缺失', '无具体工期目标', '缺少关键时间目标', '有具体工期目标', '无、没有', '高'],
    ['施工进度计划', '图表缺失', '无横道图（网络图）/没有网络图', '缺少必要图表', '有横道图和网络图', '无、没有', '高'],
    ['施工进度计划', '计划不具体', '进度计划不详细具体', '整体计划质量差', '计划详细具体', '不详细、不具体', '高'],
    ['施工技术方案', '内容缺失', '没有具体施工内容的技术方案、指标', '缺少核心技术内容', '有具体技术方案和指标', '没有、缺乏', '高'],
    ['施工技术方案', '内容偏离', '有危险化学品么？', '内容与项目不符', '内容符合项目实际', '危险化学品、不符', '高'],
    ['施工技术方案', '内容不全面', '安全和进度方面内容不够全面', '关键方面缺失', '内容全面', '不够全面', '中'],
    ['施工技术方案', '针对性不足', '针对本标段的方案论述不全面', '缺乏对本项目的针对性', '针对本标段、论述全面', '不全面、针对性', '高'],
    ['风险源管控方案', '内容不全面', '内容不够全面', '内容缺失', '内容全面', '不够全面', '中'],
    ['风险源管控方案', '针对性差', '针对性差', '缺乏针对性', '针对性强', '针对性差', '高'],
    ['风险源管控方案', '风险识别不足', '针对本项目的风险欠缺', '风险识别不全面', '风险识别全面', '欠缺、不足', '高'],
    ['关键部位质量管控措施', '内容简单', '内容比较简单', '内容过于简略', '内容详细', '比较简单', '低'],
    ['关键部位质量管控措施', '措施质量差', '关键部位质量管控措施差', '措施本身质量不高', '措施完善、有效', '差、不完善', '高'],
    ['关键部位质量管控措施', '针对性不足', '针对项目的管控欠缺', '缺乏针对性', '针对本项目', '欠缺、不足', '高'],
    ['施工进度保证措施', '内容不详细', '内容不够详细', '内容简略', '内容详细', '不够详细', '中'],
    ['质量安全文明施工保证措施', '针对性不足', '针对本项目的措施欠缺', '缺乏针对性', '针对本项目', '欠缺、不足', '高'],
    ['扬尘污染治理方案', '措施不完善', '降噪、残土排运措施差', '具体措施不到位', '措施完善、有效', '差、不完善', '高'],
    ['突发事件应急预案', '内容缺失', '（无具体原因）', '内容完全缺失或不合格', '有完善的应急预案', '无、缺失', '高'],
]

for row_data in data_analysis:
    ws1.append(row_data)

# 样式设置
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col in range(1, 8):
    cell = ws1.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

weight_colors = {'高': 'FFC7CE', '中': 'FFEB9C', '低': 'C6EFCE'}
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row in range(2, len(data_analysis) + 2):
    for col in range(1, 8):
        cell = ws1.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = data_align
        if col == 7:
            weight = cell.value
            if weight in weight_colors:
                cell.fill = PatternFill(start_color=weight_colors[weight], end_color=weight_colors[weight], fill_type='solid')
            cell.alignment = center_align

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 20
ws1.column_dimensions['C'].width = 35
ws1.column_dimensions['D'].width = 28
ws1.column_dimensions['E'].width = 28
ws1.column_dimensions['F'].width = 35
ws1.column_dimensions['G'].width = 12

ws1.row_dimensions[1].height = 35
for row in range(2, len(data_analysis) + 2):
    ws1.row_dimensions[row].height = 40

# Sheet2: 问题规律总结
ws2 = wb.create_sheet('问题规律总结')
headers2 = ['问题类型', '出现次数', '占比', '核心特征', 'AI识别策略', '提示词要点']
ws2.append(headers2)

summary_data = [
    ['针对性不足', '15', '35%', '缺乏对本项目的具体分析，使用通用模板', '检测是否包含本标段/本项目等关键词，分析内容与项目特征的匹配度', '必须要求内容针对本标段/本项目具体展开，禁止通用模板内容'],
    ['内容缺失/不全面', '12', '28%', '关键内容缺失或过于简略', '检测关键章节是否完整，内容长度是否符合要求', '要求内容完整覆盖所有评分点要求，不得遗漏关键要素'],
    ['图表缺失/错误', '8', '19%', '缺少必要的横道图、网络图或图表错误', '检测文档中是否包含必要的图表，图表内容是否正确', '必须包含横道图和网络图，图表需与文字说明一致'],
    ['内容偏离/错误', '6', '14%', '包含与项目无关的内容或基本信息错误', '检测内容是否与项目特征相符，是否存在明显错误', '内容必须与项目实际相符，基本信息（名称、范围）必须准确'],
    ['表述不清', '4', '9%', '文字晦涩、冗长或不清晰', '检测文字可读性，分析是否存在晦涩冗长的表述', '要求文字清楚、简洁明了，避免晦涩冗长的表述'],
    ['措施质量差', '3', '7%', '措施本身不完善或效果差', '评估措施的可行性和完善程度', '要求措施具体、可执行、有效'],
]

for row_data in summary_data:
    ws2.append(row_data)

for col in range(1, 7):
    cell = ws2.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

for row in range(2, len(summary_data) + 2):
    for col in range(1, 7):
        cell = ws2.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = data_align

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 35
ws2.column_dimensions['E'].width = 45
ws2.column_dimensions['F'].width = 45

ws2.row_dimensions[1].height = 35
for row in range(2, len(summary_data) + 2):
    ws2.row_dimensions[row].height = 50

# Sheet3: AI智能辅助评标提示词
ws3 = wb.create_sheet('AI评标提示词')
headers3 = ['评分点', '1分判定提示词', '0分判定提示词', '关键检查项']
ws3.append(headers3)

prompt_data = [
    ['综合阐述', '检查是否对本项目施工进行了综合阐述，文字是否清楚，目标是否明确，内容是否针对本标段/本项目。', '如果发现以下情况之一，判定为0分：1）目标不清晰；2）内容偏离（如未提及关键改造内容）；3）包含与项目无关的内容；4）范围错误；5）完全缺乏针对性内容。', '目标明确性、内容针对性、范围准确性、文字清晰度'],
    ['管理架构', '检查组织架构图是否正确，岗位配备是否齐全，职责描述是否明确具体，是否针对本标段。', '如果发现以下情况之一，判定为0分：1）组织架构图错误；2）职责描述不具体、不全面；3）缺乏对本标段的针对性安排。', '架构正确性、职责明确性、内容全面性、针对性'],
    ['施工进度计划', '检查是否有符合工期要求的进度计划，文字表述是否与横道图（网络图）一致，阶段划分是否清晰，时间节点是否明确。', '如果发现以下情况之一，判定为0分：1）无横道图或网络图；2）描述不详细、不具体；3）阶段划分不清，未标注日历天数；4）无具体工期目标。', '图表完整性、描述一致性、阶段清晰度、时间明确性'],
    ['施工技术方案', '检查是否编写了针对本标段施工内容的技术方案，是否确保施工安全、质量及进度要求，内容是否全面。', '如果发现以下情况之一，判定为0分：1）没有具体施工内容的技术方案；2）内容与项目不符；3）安全、进度方面内容不全面；4）针对性不足。', '内容完整性、项目符合性、安全质量进度覆盖、针对性'],
    ['风险源管控方案', '检查是否针对标段内容编制了主要风险源管控方案，内容是否全面，针对性是否强。', '如果发现以下情况之一，判定为0分：1）内容不够全面；2）针对性差；3）针对本项目的风险识别欠缺。', '风险识别全面性、管控措施针对性、内容完整性'],
    ['关键部位质量管控措施', '检查是否提供了关键部位质量管控措施，措施是否完善、有效，是否针对本项目。', '如果发现以下情况之一，判定为0分：1）内容过于简单；2）措施质量差；3）针对本项目的管控欠缺。', '措施完善性、有效性、针对性'],
    ['施工进度保证措施', '检查是否有满足竣工工期要求的施工进度保证措施，内容是否详细。', '如果发现以下情况之一，判定为0分：1）内容不够详细、不具体。', '措施详细性、可行性'],
    ['质量安全文明施工保证措施', '检查是否有符合本标段项目实际情况的质量、安全、文明施工保证措施，是否针对本项目。', '如果发现以下情况之一，判定为0分：1）针对本项目的措施欠缺。', '措施针对性、符合性'],
    ['扬尘污染治理方案', '检查是否编制了扬尘污染源治理方案，包括防尘、降噪、残土排运等措施，措施是否完善。', '如果发现以下情况之一，判定为0分：1）降噪、残土排运等措施差。', '措施完整性、有效性'],
    ['突发事件应急预案', '检查是否编制了突发事件应急预案，包括组织措施及紧急情况的抢救措施等。', '如果发现以下情况之一，判定为0分：1）内容缺失；2）不符合要求。', '预案完整性、组织措施、抢救措施'],
]

for row_data in prompt_data:
    ws3.append(row_data)

for col in range(1, 5):
    cell = ws3.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

for row in range(2, len(prompt_data) + 2):
    for col in range(1, 5):
        cell = ws3.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = data_align

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 50
ws3.column_dimensions['C'].width = 50
ws3.column_dimensions['D'].width = 35

ws3.row_dimensions[1].height = 35
for row in range(2, len(prompt_data) + 2):
    ws3.row_dimensions[row].height = 80

# Sheet4: AI评标规则库
ws4 = wb.create_sheet('AI评标规则库')
headers4 = ['规则ID', '规则名称', '适用评分点', '触发条件', '判定结果', '优先级']
ws4.append(headers4)

rules_data = [
    ['R001', '针对性检测', '全部', '内容未包含"本标段"、"本项目"、"本小区"等关键词', '0分', '高'],
    ['R002', '内容偏离检测', '综合阐述', '包含融资、股权、抵押、气瓶、危险化学品等与项目无关的内容', '0分', '高'],
    ['R003', '基本信息错误', '综合阐述', '项目名称、范围与招标文件不符', '0分', '高'],
    ['R004', '图表缺失检测', '施工进度计划', '未包含横道图或网络图', '0分', '高'],
    ['R005', '图表说明不足', '施工进度计划', '有横道图但无详细文字说明', '0分', '中'],
    ['R006', '时间节点缺失', '施工进度计划', '未标注日历天数或工期目标', '0分', '高'],
    ['R007', '技术方案缺失', '施工技术方案', '无具体施工内容的技术方案', '0分', '高'],
    ['R008', '风险识别不足', '风险源管控方案', '未针对本项目识别具体风险', '0分', '高'],
    ['R009', '措施质量检测', '关键部位质量管控措施', '措施描述过于简单或质量差', '0分', '高'],
    ['R010', '措施针对性', '质量安全文明施工保证措施', '未针对本项目编制措施', '0分', '高'],
    ['R011', '环保措施检测', '扬尘污染治理方案', '降噪、残土排运等措施不完善', '0分', '高'],
    ['R012', '内容完整性', '全部', '关键内容缺失或章节不完整', '0分', '高'],
]

for row_data in rules_data:
    ws4.append(row_data)

for col in range(1, 7):
    cell = ws4.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align

for row in range(2, len(rules_data) + 2):
    for col in range(1, 7):
        cell = ws4.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = data_align

ws4.column_dimensions['A'].width = 12
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 50
ws4.column_dimensions['E'].width = 12
ws4.column_dimensions['F'].width = 12

ws4.row_dimensions[1].height = 35
for row in range(2, len(rules_data) + 2):
    ws4.row_dimensions[row].height = 45

# 保存文件
wb.save('技术标0分问题深度分析及AI评标规则库.xlsx')
print('Excel文件已创建: 技术标0分问题深度分析及AI评标规则库.xlsx')