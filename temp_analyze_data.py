import pandas as pd

# 读取数据
file_path = 'D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\temp\\专区低收益统计结果.xlsx'
zhibiao1 = pd.read_excel(file_path, sheet_name='指标一-超1年总收益<10w')
zhibiao2 = pd.read_excel(file_path, sheet_name='指标二-25年前25年收益<10w')
zhibiao3 = pd.read_excel(file_path, sheet_name='指标三-超1年总收益<5w')
zhibiao4 = pd.read_excel(file_path, sheet_name='指标四-25年前25年收益<5w')
zhibiao5 = pd.read_excel(file_path, sheet_name='指标五-26年产生收益')
zhibiao6 = pd.read_excel(file_path, sheet_name='指标六-25年有收益26年无')

# 用专区号作为唯一标识
zhibiao1_set = set(zhibiao1['专区号'].dropna())
zhibiao2_set = set(zhibiao2['专区号'].dropna())
zhibiao3_set = set(zhibiao3['专区号'].dropna())
zhibiao4_set = set(zhibiao4['专区号'].dropna())
zhibiao5_set = set(zhibiao5['专区号'].dropna())
zhibiao6_set = set(zhibiao6['专区号'].dropna())

print('=== 数据关系分析 ===')
print(f'指标一（超1年总收益<10w）: {len(zhibiao1_set)}个专区')
print(f'指标二（25年前25年收益<10w）: {len(zhibiao2_set)}个专区')
print(f'指标三（超1年总收益<5w）: {len(zhibiao3_set)}个专区')
print(f'指标四（25年前25年收益<5w）: {len(zhibiao4_set)}个专区')
print(f'指标五（26年产生收益）: {len(zhibiao5_set)}个专区')
print(f'指标六（25年有收益26年无）: {len(zhibiao6_set)}个专区')

# 分析包含关系
print('\n=== 包含关系验证 ===')
print(f'指标三是否属于指标一的子集: {zhibiao3_set.issubset(zhibiao1_set)}')
print(f'指标四是否属于指标二的子集: {zhibiao4_set.issubset(zhibiao2_set)}')

# 计算真正的独立数量
# 指标一独有的（<10w但≥5w）
zhibiao1_only = zhibiao1_set - zhibiao3_set
print(f'\n指标一独有的（5w<=收益<10w）: {len(zhibiao1_only)}个')

# 指标二独有的（<10w但≥5w）
zhibiao2_only = zhibiao2_set - zhibiao4_set
print(f'指标二独有的（5w<=25年收益<10w）: {len(zhibiao2_only)}个')

# 指标一和指标二的重叠
zhibiao1_and_2 = zhibiao1_set & zhibiao2_set
print(f'指标一和指标二重叠: {len(zhibiao1_and_2)}个')

# 指标三和指标四的重叠
zhibiao3_and_4 = zhibiao3_set & zhibiao4_set
print(f'指标三和指标四重叠: {len(zhibiao3_and_4)}个')

# 重新分类统计
print('\n=== 重新分类统计 ===')
print(f'超1年且总收益5-10w: {len(zhibiao1_only)}个（指标一独有）')
print(f'超1年且总收益<5w: {len(zhibiao3_set)}个（指标三）')
print(f'25年前且25年收益5-10w: {len(zhibiao2_only)}个（指标二独有）')
print(f'25年前且25年收益<5w: {len(zhibiao4_set)}个（指标四）')

# 计算不重复的总数
print('\n=== 不重复统计 ===')
all_zones = zhibiao1_set | zhibiao2_set | zhibiao3_set | zhibiao4_set | zhibiao5_set | zhibiao6_set
print(f'所有指标涉及的不重复专区总数: {len(all_zones)}个')

# 按风险等级重新统计
print('\n=== 按风险等级重新统计 ===')
# 红色：超1年且总收益<10w（指标一全部）
red_zones = zhibiao1_set
print(f'红色-重点关注（超1年且总收益<10w）: {len(red_zones)}个')

# 橙色：25年前接入且25年收益<10w，但不包含在红色中的
orange_zones = zhibiao2_set - red_zones
print(f'橙色-需改进（25年前且25年收益<10w，排除红色）: {len(orange_zones)}个')

# 灰色：25年有收益但26年无，且不包含在红橙中的
grey_zones = zhibiao6_set - red_zones - orange_zones
print(f'灰色-流失风险（25年有但26年无，排除红橙）: {len(grey_zones)}个')

# 黄色：26年有收益，且不包含在红橙灰中的
yellow_zones = zhibiao5_set - red_zones - orange_zones - grey_zones
print(f'黄色-观察（26年有收益，排除红橙灰）: {len(yellow_zones)}个')

# 保存分析结果到Excel
print('\n=== 保存分析结果 ===')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# 样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 工作表1：重新分类统计
ws1 = wb.create_sheet('重新分类统计')
data1 = [
    ['风险等级', '条件', '数量', '说明'],
    ['红色-重点关注', '超1年且总收益<10w', len(red_zones), '指标一全部'],
    ['橙色-需改进', '25年前且25年收益<10w（排除红色）', len(orange_zones), '指标二独有'],
    ['灰色-流失风险', '25年有但26年无（排除红橙）', len(grey_zones), '指标六排除重叠'],
    ['黄色-观察', '26年有收益（排除红橙灰）', len(yellow_zones), '指标五排除重叠'],
    ['合计', '不重复专区总数', len(all_zones), '去重后总数'],
]
for r_idx, row in enumerate(data1, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 40
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 20

# 工作表2：详细分类
ws2 = wb.create_sheet('详细分类')
data2 = [
    ['分类', '数量', '说明'],
    ['超1年且总收益5-10w', len(zhibiao1_only), '指标一独有'],
    ['超1年且总收益<5w', len(zhibiao3_set), '指标三'],
    ['25年前且25年收益5-10w', len(zhibiao2_only), '指标二独有'],
    ['25年前且25年收益<5w', len(zhibiao4_set), '指标四'],
    ['26年产生收益', len(zhibiao5_set), '指标五'],
    ['25年有收益但26年无', len(zhibiao6_set), '指标六'],
]
for r_idx, row in enumerate(data2, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 20

# 保存
output_path = 'D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\temp\\数据分析结果_修正版.xlsx'
wb.save(output_path)
print(f'分析结果已保存: {output_path}')
