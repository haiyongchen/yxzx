import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

base_dir = r'D:\work\运营中心\yxzx\阳光优采\需求\在线支付'

# Find files
docx_path = None
xlsx_path = None
for f in os.listdir(base_dir):
    if f.endswith('.docx'):
        docx_path = os.path.join(base_dir, f)
    elif f.endswith('.xlsx'):
        xlsx_path = os.path.join(base_dir, f)

# Read DOCX using python-docx
print("=== 阳光优采-在线支付需求优化.docx ===\n")
try:
    from docx import Document
    doc = Document(docx_path)
    for para in doc.paragraphs:
        if para.text.strip():
            print(para.text)
except Exception as e:
    print(f"Error reading docx: {e}")

print("\n\n=== 阳光优采在线支付功能业务改造.xlsx ===\n")
# Read XLSX using openpyxl
try:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            row_data = [str(cell) if cell is not None else '' for cell in row]
            if any(cell.strip() for cell in row_data):
                print('\t'.join(row_data))
except Exception as e:
    print(f"Error reading xlsx: {e}")
