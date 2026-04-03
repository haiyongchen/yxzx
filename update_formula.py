import openpyxl

# 加载Excel文件
wb = openpyxl.load_workbook('D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\专区信息汇总表_按省份分类.xlsx')

# 新的公式模板
p_formula = '=IF(O{row}="","",IF(DATEDIF(O{row},NOW(),"Y")>0,DATEDIF(O{row},NOW(),"Y")&"年","")&IF(DATEDIF(O{row},NOW(),"YM")>0,DATEDIF(O{row},NOW(),"YM")&"月","")&DATEDIF(O{row},NOW(),"MD")&"天")'
r_formula = '=IF(Q{row}="","",IF(DATEDIF(Q{row},NOW(),"Y")>0,DATEDIF(Q{row},NOW(),"Y")&"年","")&IF(DATEDIF(Q{row},NOW(),"YM")>0,DATEDIF(Q{row},NOW(),"YM")&"月","")&DATEDIF(Q{row},NOW(),"MD")&"天")'

# 遍历所有sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f'处理Sheet: {sheet_name}')
    
    # 获取数据行数
    max_row = sheet.max_row
    
    # 从第2行开始设置公式（假设第1行是表头）
    for row in range(2, max_row + 1):
        # P列: 计算O列距离当前时间的年月日
        sheet[f'P{row}'] = p_formula.format(row=row)
        
        # R列: 计算Q列距离当前时间的年月日
        sheet[f'R{row}'] = r_formula.format(row=row)
    
    print(f'  已更新P列和R列公式，共{max_row-1}行数据')

# 保存文件
wb.save('D:\\work\\运营中心\\yxzx\\新点e交易相关材料\\日常数据运维工具\\专区信息汇总表_按省份分类.xlsx')
print('\n文件保存成功！')
