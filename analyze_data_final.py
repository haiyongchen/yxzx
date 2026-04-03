#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完整数据分析 - 包含新指标
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 读取原始数据
file_path = r'D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\temp\专区信息汇总表_中原华北_合并.xlsx'
df = pd.read_excel(file_path)

# 获取关键列
confirm_time = df.iloc[:, 14]  # 确认接入时间
total_revenue = df.iloc[:, 21] # 总收益
revenue_25 = df.iloc[:, 19]    # 25年总收益
revenue_26 = df.iloc[:, 20]    # 26年总收益
total_cost = df.iloc[:, 25]    # 总成本

# 当前日期
current_date = datetime(2026, 4, 2)
one_year_ago = datetime(2025, 4, 2)
year_2025_start = datetime(2025, 1, 1)

# 清理数据
confirm_time_clean = pd.to_datetime(confirm_time, errors='coerce')
total_revenue_clean = pd.to_numeric(total_revenue, errors='coerce')
revenue_25_clean = pd.to_numeric(revenue_25, errors='coerce')
revenue_26_clean = pd.to_numeric(revenue_26, errors='coerce')

# 原有指标
condition1 = (confirm_time_clean < one_year_ago) & (total_revenue_clean < 100000) & (total_revenue_clean.notna())
condition2 = (confirm_time_clean < year_2025_start) & (revenue_25_clean < 100000) & (revenue_25_clean.notna())
condition3 = (confirm_time_clean < one_year_ago) & (total_revenue_clean < 50000) & (total_revenue_clean.notna())
condition4 = (confirm_time_clean < year_2025_start) & (revenue_25_clean < 50000) & (revenue_25_clean.notna())
condition5 = (revenue_26_clean > 0) & (revenue_26_clean.notna())
condition6 = (revenue_25_clean > 0) & (revenue_25_clean.notna()) & ((revenue_26_clean == 0) | (revenue_26_clean.isna()))

# 新指标七：接入超过1年，总收益为0
condition7 = (confirm_time_clean < one_year_ago) & ((total_revenue_clean == 0) | (total_revenue_clean.isna()))

# 获取各指标数据
zhibiao1 = df[condition1]
zhibiao2 = df[condition2]
zhibiao3 = df[condition3]
zhibiao4 = df[condition4]
zhibiao5 = df[condition5]
zhibiao6 = df[condition6]
zhibiao7 = df[condition7]

# 用专区号作为唯一标识
zhibiao1_set = set(zhibiao1.iloc[:, 1].dropna())
zhibiao2_set = set(zhibiao2.iloc[:, 1].dropna())
zhibiao3_set = set(zhibiao3.iloc[:, 1].dropna())
zhibiao4_set = set(zhibiao4.iloc[:, 1].dropna())
zhibiao5_set = set(zhibiao5.iloc[:, 1].dropna())
zhibiao6_set = set(zhibiao6.iloc[:, 1].dropna())
zhibiao7_set = set(zhibiao7.iloc[:, 1].dropna())

print('=== 七大指标原始统计 ===')
print(f'指标一（超1年总收益<10w）: {len(zhibiao1_set)}个')
print(f'指标二（25年前25年收益<10w）: {len(zhibiao2_set)}个')
print(f'指标三（超1年总收益<5w）: {len(zhibiao3_set)}个')
print(f'指标四（25年前25年收益<5w）: {len(zhibiao4_set)}个')
print(f'指标五（26年产生收益）: {len(zhibiao5_set)}个')
print(f'指标六（25年有收益26年无）: {len(zhibiao6_set)}个')
print(f'指标七（超1年总收益为0）: {len(zhibiao7_set)}个')

# 分析包含关系
print('\n=== 包含关系分析 ===')
print(f'指标三是否属于指标一的子集: {zhibiao3_set.issubset(zhibiao1_set)}')
print(f'指标四是否属于指标二的子集: {zhibiao4_set.issubset(zhibiao2_set)}')
print(f'指标七是否属于指标一的子集: {zhibiao7_set.issubset(zhibiao1_set)}')

# 指标一和指标七的关系（指标七应该是指标一的一部分）
zhibiao7_in_zhibiao1 = zhibiao7_set.intersection(zhibiao1_set)
print(f'指标七在指标一中的数量: {len(zhibiao7_in_zhibiao1)} / {len(zhibiao7_set)}')

# 计算独立数量
zhibiao1_only = zhibiao1_set - zhibiao3_set - zhibiao7_set  # 指标一独有的（5-10w，不含0）
zhibiao1_has_revenue = zhibiao1_set - zhibiao7_set  # 指标一有收益的（不含0）

print(f'\n指标一独有的（5w<=收益<10w，不含0）: {len(zhibiao1_only)}个')
print(f'指标一（有收益的）: {len(zhibiao1_has_revenue)}个')
print(f'指标三（<5w）: {len(zhibiao3_set)}个')
print(f'指标七（=0）: {len(zhibiao7_set)}个')

# 验证：指标一 = 指标一独有 + 指标三 + 指标七
print(f'\n验证: {len(zhibiao1_only)} + {len(zhibiao3_set)} + {len(zhibiao7_set)} = {len(zhibiao1_only) + len(zhibiao3_set) + len(zhibiao7_set)} (应为{len(zhibiao1_set)})')

# 去重后的风险等级统计
print('\n=== 去重后的风险等级统计 ===')

# 红色：超1年且总收益<10w（指标一全部）
red_zones = zhibiao1_set
print(f'红色-重点关注（超1年且总收益<10w）: {len(red_zones)}个')

# 橙色：25年前且25年收益<10w，但不包含在红色中的
orange_zones = zhibiao2_set - red_zones
print(f'橙色-需改进（25年前且25年收益<10w，排除红色）: {len(orange_zones)}个')

# 灰色：25年有收益但26年无，且不包含在红橙中的
grey_zones = zhibiao6_set - red_zones - orange_zones
print(f'灰色-流失风险（25年有但26年无，排除红橙）: {len(grey_zones)}个')

# 黄色：26年有收益，且不包含在红橙灰中的
yellow_zones = zhibiao5_set - red_zones - orange_zones - grey_zones
print(f'黄色-观察（26年有收益，排除红橙灰）: {len(yellow_zones)}个')

# 黑色（新增）：超1年且总收益为0
black_zones = zhibiao7_set
print(f'黑色-零收益（超1年且总收益为0）: {len(black_zones)}个')

# 注意：黑色应该是红色的一部分
black_in_red = black_zones.intersection(red_zones)
print(f'黑色在红色中的数量: {len(black_in_red)} / {len(black_zones)}')

# 不重复的总数（红+橙+灰+黄，黑已经包含在红中）
all_zones = red_zones | orange_zones | grey_zones | yellow_zones
print(f'\n不重复专区总数: {len(all_zones)}个')

# 保存到Excel
print('\n=== 保存分析结果 ===')
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# 样式
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 工作表1：七大指标原始统计
ws1 = wb.create_sheet('七大指标原始统计')
data1 = [
    ['指标', '条件', '数量', '包含关系', '备注'],
    ['指标一', '超1年且总收益<10w', len(zhibiao1_set), '包含指标三、指标七', ''],
    ['指标二', '25年前且25年收益<10w', len(zhibiao2_set), '包含指标四', ''],
    ['指标三', '超1年且总收益<5w', len(zhibiao3_set), '属于指标一', '子集'],
    ['指标四', '25年前且25年收益<5w', len(zhibiao4_set), '属于指标二', '子集'],
    ['指标五', '26年产生收益', len(zhibiao5_set), '-', ''],
    ['指标六', '25年有收益但26年无', len(zhibiao6_set), '-', ''],
    ['指标七（新增）', '超1年且总收益为0', len(zhibiao7_set), '属于指标一', '新增'],
]
for r_idx, row in enumerate(data1, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 35
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 15

# 工作表2：去重后风险等级
ws2 = wb.create_sheet('去重后风险等级')
data2 = [
    ['风险等级', '条件', '数量', '占比', '说明'],
    ['🔴 红色-重点关注', '超1年且总收益<10w', len(red_zones), f'{len(red_zones)/len(all_zones)*100:.1f}%', '指标一全部'],
    ['⚫ 黑色-零收益', '超1年且总收益为0', len(black_zones), f'{len(black_zones)/len(all_zones)*100:.1f}%', '指标七（红色子集）'],
    ['🟠 橙色-需改进', '25年前且25年收益<10w（排除红色）', len(orange_zones), f'{len(orange_zones)/len(all_zones)*100:.1f}%', '指标二独有'],
    ['⚪ 灰色-流失风险', '25年有但26年无（排除红橙）', len(grey_zones), f'{len(grey_zones)/len(all_zones)*100:.1f}%', '指标六排除重叠'],
    ['🟡 黄色-观察', '26年有收益（排除红橙灰）', len(yellow_zones), f'{len(yellow_zones)/len(all_zones)*100:.1f}%', '指标五排除重叠'],
    ['合计', '不重复专区总数', len(all_zones), '100%', '去重后总数'],
]
for r_idx, row in enumerate(data2, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 25

# 工作表3：指标七详细数据
ws3 = wb.create_sheet('指标七-零收益专区')
result7 = zhibiao7[[df.columns[0], df.columns[1], df.columns[2], df.columns[14], df.columns[21]]].copy()
result7.columns = ['合同编号', '专区号', '专区名称', '确认接入时间', '总收益']
for r_idx, row in enumerate([['合同编号', '专区号', '专区名称', '确认接入时间', '总收益']] + result7.values.tolist(), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws3.cell(row=r_idx, column=c_idx, value=value)
        cell.border = border
        if r_idx == 1:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 30
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 12

# 保存
output_path = r'D:\work\运营中心\yxzx\新点e交易相关材料\日常数据运维工具\temp\数据分析结果_完整版.xlsx'
wb.save(output_path)
print(f'\n分析结果已保存: {output_path}')